# -*- coding: utf-8 -*-
"""Terminology database manager for Bach Cantata translation project.

Maintains a shared Excel glossary at the workspace root level:
  巴赫康塔塔术语库.xlsx

Columns: 原文术语 (德语), 译文术语 (中文和合本), 所属康塔塔编号, 出现频次, 备注说明

After each cantata translation, this module scans the newly generated
glossary.json and updates the shared Excel:
  - New terms → appended
  - Existing terms → frequency incremented, BWV number appended
  - Translation differences across cantatas → noted in 备注
"""

import os

from .logger import get_logger

log = get_logger()

# ── Constants ──────────────────────────────────────────────────────────
WORKSPACE_ROOT = os.path.normpath(
    os.path.join(os.path.dirname(__file__), '..')
)
DB_FILENAME = '巴赫康塔塔术语库.xlsx'
DB_PATH = os.path.join(WORKSPACE_ROOT, DB_FILENAME)

HEADERS = ['原文术语 (德语)', '译文术语 (中文和合本)',
           '所属康塔塔编号', '出现频次', '备注说明']


def _ensure_db_exists():
    """Create a blank terminology database if one doesn't exist."""
    if os.path.exists(DB_PATH):
        return
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '巴赫康塔塔术语库'

    header_font = Font(name='宋体', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(
        start_color='1A1A2E', end_color='1A1A2E', fill_type='solid'
    )
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    col_widths = [26, 24, 22, 10, 50]
    for i, (h, w) in enumerate(zip(HEADERS, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = 'A2'
    wb.save(DB_PATH)
    log.info(f"[glossary_db] Created blank database: {DB_PATH}")


def _read_db():
    """Read existing Excel into a dict keyed by lowercased German term."""
    if not os.path.exists(DB_PATH):
        _ensure_db_exists()
        return {}

    from openpyxl import load_workbook
    wb = load_workbook(DB_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return {}

    db = {}
    for row in rows[1:]:
        if not row[0]:
            continue
        key = str(row[0]).strip().lower()
        db[key] = {
            'german': str(row[0]).strip(),
            'chinese': str(row[1]).strip() if row[1] else '',
            'bwvs': str(row[2]).strip() if row[2] else '',
            'count': int(row[3]) if row[3] else 0,
            'notes': str(row[4]).strip() if row[4] else '',
            'excel_row': None,  # filled on write-back
        }
    return db


def _write_db(records):
    """Write all records sorted alphabetically to the Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    sorted_recs = sorted(records, key=lambda r: r['german'].lower())

    wb = Workbook()
    ws = wb.active
    ws.title = '巴赫康塔塔术语库'

    header_font = Font(name='宋体', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(
        start_color='1A1A2E', end_color='1A1A2E', fill_type='solid'
    )
    cell_font = Font(name='宋体', size=10)
    cell_font_de = Font(name='Times New Roman', size=10)
    note_font = Font(name='宋体', size=9, color='555555')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    wrap_align = Alignment(wrap_text=True, vertical='center')
    center_align = Alignment(horizontal='center', vertical='center')

    col_widths = [26, 24, 22, 10, 50]
    for i, (h, w) in enumerate(zip(HEADERS, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 28

    for ri, rec in enumerate(sorted_recs, 2):
        c1 = ws.cell(row=ri, column=1, value=rec['german'])
        c1.font = cell_font_de
        c1.border = thin_border
        c1.alignment = wrap_align

        c2 = ws.cell(row=ri, column=2, value=rec['chinese'])
        c2.font = cell_font
        c2.border = thin_border
        c2.alignment = wrap_align

        c3 = ws.cell(row=ri, column=3, value=rec['bwvs'])
        c3.font = cell_font
        c3.border = thin_border
        c3.alignment = center_align

        c4 = ws.cell(row=ri, column=4, value=rec['count'])
        c4.font = cell_font
        c4.border = thin_border
        c4.alignment = center_align

        c5 = ws.cell(row=ri, column=5, value=rec.get('notes', ''))
        c5.font = note_font
        c5.border = thin_border
        c5.alignment = wrap_align

        ws.row_dimensions[ri].height = 22

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:E{ri}'
    wb.save(DB_PATH)


def update_from_glossary(bwv_number, glossary_entries):
    """Merge new glossary entries into the shared terminology database.

    Args:
        bwv_number: int or str — the BWV number processed
        glossary_entries: list of dicts with keys:
            german, chinese_cuv, note

    Each entry is merged by lowercased German term:
      - New term → appended with count=1, bwvs=current BWV
      - Existing same translation → frequency +1, BWV added
      - Existing different translation → frequency +1, BWV added,
        and discrepancy noted in 备注
    """
    bwv_label = f'BWV {bwv_number}'

    _ensure_db_exists()
    db = _read_db()

    new_terms = 0
    updated_terms = 0
    translation_conflicts = 0

    for entry in glossary_entries:
        key = entry['german'].strip().lower()
        chinese = entry.get('chinese_cuv', '').strip()
        note = entry.get('note', '').strip()

        if not key:
            continue

        if key in db:
            # Existing term
            rec = db[key]

            # Check if this BWV is already counted
            existing_bwvs = set(b.strip() for b in rec['bwvs'].split(',') if b.strip())
            is_new_bwv = bwv_label not in existing_bwvs

            if is_new_bwv:
                rec['count'] += 1

            # Update BWV list
            existing_bwvs.add(bwv_label)
            rec['bwvs'] = ', '.join(sorted(existing_bwvs, key=lambda x: (x.split()[-1].isdigit() and int(x.split()[-1])) or float('inf')))

            # Check for translation differences
            if chinese and chinese != rec['chinese']:
                # Append alternative if not already listed
                if f'「{chinese}」' not in rec['chinese']:
                    if ' (另见: ' not in rec['chinese']:
                        rec['chinese'] = f"{rec['chinese']} (另见: {chinese})"
                    else:
                        # Multiple alternatives already exist
                        if f'「{chinese}」' not in rec['chinese']:
                            rec['chinese'] = rec['chinese'].rstrip(')') + f', {chinese})'

                # Add discrepancy to notes
                disc_note = f'{bwv_label}译为「{chinese}」'
                if '不同康塔塔译法差异' not in rec['notes']:
                    rec['notes'] = f'不同康塔塔译法差异: {disc_note}\n{rec["notes"]}'
                elif disc_note not in rec['notes']:
                    rec['notes'] = rec['notes'].replace(
                        '不同康塔塔译法差异: ',
                        f'不同康塔塔译法差异: {disc_note}; '
                    )
                translation_conflicts += 1

            # Append note if new
            if note and note not in rec['notes']:
                if rec['notes']:
                    rec['notes'] += f'; {note}'
                else:
                    rec['notes'] = note

            updated_terms += 1
        else:
            # New term
            db[key] = {
                'german': entry['german'].strip(),
                'chinese': chinese,
                'bwvs': bwv_label,
                'count': 1,
                'notes': note,
            }
            new_terms += 1

    # Write back
    _write_db(list(db.values()))

    log.info(
        f"[glossary_db] Updated term DB: "
        f"{new_terms} new, {updated_terms} updated, "
        f"{translation_conflicts} translation conflicts"
    )
    return {
        'new': new_terms,
        'updated': updated_terms,
        'conflicts': translation_conflicts,
        'total': len(db),
    }


def load_terms():
    """Load the full terminology database as a list of dicts.

    Returns:
        list of dicts with keys: german, chinese, bwvs, count, notes
    """
    _ensure_db_exists()
    db = _read_db()
    return list(db.values())
